import openpyxl, re, os
REF="/Users/wonda/Documents/GitHub/CoCASUCSB/reference"
OUT="/Users/wonda/Documents/GitHub/CoCASUCSB/supabase/seed.sql"
def load(name): return openpyxl.load_workbook(os.path.join(REF,name),data_only=True,read_only=True)
def g(r,i): return r[i] if (i is not None and i < len(r)) else None
def q(s):
    if s is None: return "null"
    s=str(s).strip()
    return "null" if s=="" else "'"+s.replace("'","''")+"'"
def num(v):
    if v is None or v=="" : return "null"
    try: return str(round(float(v),2))
    except:
        m=re.search(r'-?\d[\d,]*\.?\d*',str(v))
        return str(float(m.group().replace(',',''))) if m else "null"
STOP={'committee','board','commission','commissions','unit','units','office','of','the','as','a','for','and'}
def norm(name):
    s=str(name).lower().replace('&',' and ')
    s=re.sub(r'[^a-z0-9 ]',' ',s)
    return " ".join(t for t in s.split() if t and t not in STOP)
def slugify(name):
    s=re.sub(r'[^a-z0-9]+','-',str(name).lower()).strip('-'); return s[:60] or 'bcu'
def acronym(name):
    w=[x for x in re.sub(r'[^A-Za-z0-9 ]',' ',str(name)).split() if x]
    return "".join(x[0] for x in w).upper()[:8]
bcus={}; usedslugs=set()
def add_bcu(name, short=None, website=None, cname=None, cemail=None, btype='committee'):
    if not name or str(name).strip()=="" : return None
    name=str(name).strip(); k=norm(name)
    if not k: return None
    if k not in bcus:
        sl=slugify(short or name)
        while sl in usedslugs: sl=sl+"-x"
        usedslugs.add(sl)
        bcus[k]=dict(name=name,short=(short or acronym(name)),slug=sl,type=btype,
                     website=website,cname=cname,cemail=cemail)
    else:
        b=bcus[k]
        if len(name)>len(b['name']): b['name']=name
        for f,val in (('website',website),('cname',cname),('cemail',cemail)):
            if val and not b.get(f): b[f]=val
        if short and (not b['short'] or len(short)<len(b['short'])): b['short']=short
    return k
persons={}; appts=[]
def add_person(full_name, ucsb=None, asem=None):
    full_name=str(full_name).strip(); key=(ucsb or full_name).strip().lower()
    if key not in persons: persons[key]=dict(full_name=full_name,ucsb=ucsb,asem=asem)
    else:
        p=persons[key]
        if asem and not p['asem']: p['asem']=asem
        if ucsb and not p['ucsb']: p['ucsb']=ucsb
    return key
positions=[]; budget=[]

# 1. Recruitment
wb=load("26-27 OPEN AS Positions Recruitment.xlsx"); ws=wb["Recruitment"]
rows=list(ws.iter_rows(values_only=True))
hdr=[(str(c).strip() if c else "") for c in rows[0]]
def col(*keys):
    for i,h in enumerate(hdr):
        for k in keys:
            if k.lower() in h.lower(): return i
    return None
ci_cn=col("name of contact"); ci_em=col("email"); ci_cond=col("conduct")
ci_adv=col("advertis"); ci_web=col("committee website","website"); ci_notes=col("notes")
pos_cols=[i for i,h in enumerate(hdr) if h.lower().startswith("position #")]
SKIP=re.compile(r'^(all positions|n/?a|none|tbd|x)\b',re.I)
for r in rows[1:]:
    bcu_name=g(r,0)
    if not bcu_name: continue
    bn=str(bcu_name).strip()
    if bn.lower().startswith(("example","avove","above")) or "positions above" in bn.lower(): continue
    web=g(r,ci_web)
    if web and ("no website" in str(web).lower() or str(web).strip().lower()=="n/a"): web=None
    cond=str(g(r,ci_cond) or "").strip().lower()
    adv=str(g(r,ci_adv) or "").strip()
    notes=str(g(r,ci_notes) or "").strip()
    k=add_bcu(bn, website=(str(web).strip() if web else None),
              cname=(str(g(r,ci_cn)).strip() if g(r,ci_cn) else None),
              cemail=(str(g(r,ci_em)).strip() if g(r,ci_em) else None))
    if not k: continue
    nlow=notes.lower()
    if cond.startswith("yes"): routing="coc_interview"
    elif "forward" in nlow: routing="forward_to_bcu"
    elif "external" in nlow or "google form" in nlow or "flyer" in nlow: routing="external_form"
    else: routing="unknown"
    coc_adv=bool(adv) and not adv.lower().startswith("no")
    ext=str(web).strip() if (web and routing in ("external_form","unknown")) else None
    for pc in pos_cols:
        cell=g(r,pc)
        if not cell: continue
        title=re.sub(r'\s+',' ',str(cell).strip())[:160]
        if not title or SKIP.match(title): continue
        positions.append(dict(bcu=k,title=title,status="open",routing=routing,
                              external_url=ext,coc_adv=coc_adv,notes=notes or None))

# 2. CoC board
coc_k=add_bcu("Committee on Committees",short="CoC",website="https://coc.as.ucsb.edu")
ws=load("CoC Board Contact Sheet 2025-2026.xlsx").worksheets[0]
crows=list(ws.iter_rows(values_only=True))
for r in crows[4:]:
    posn=g(r,2)
    if not posn: continue
    full=f"{g(r,0) or ''} {g(r,1) or ''}".strip()
    if not full: continue
    pk=add_person(full, ucsb=(str(g(r,3)).strip() if g(r,3) else None),
                  asem=(str(g(r,4)).strip() if g(r,4) else None))
    role=str(posn).strip()
    appts.append(dict(pk=pk,bcu=coc_k,role=role,is_chair=("chair" in role.lower() and "vice" not in role.lower()),term="2025-26"))

# 3. AS Chairs roster
ws=load("AS Chairs (for mailing list).xlsx").worksheets[0]
arows=list(ws.iter_rows(values_only=True))
ah=[(str(c).strip().lower() if c else "") for c in arows[0]]
def acol(*keys):
    for i,h in enumerate(ah):
        for kk in keys:
            if kk in h: return i
    return None
i_org=acol("org"); i_pos=acol("position"); i_name=acol("name"); i_ucsb=acol("ucsb email"); i_asem=acol("as email")
for r in arows[1:]:
    org=g(r,i_org); nm=g(r,i_name)
    if not org or not nm: continue
    bk=add_bcu(str(org).strip())
    pk=add_person(str(nm).strip(),
                  ucsb=(str(g(r,i_ucsb)).strip() if g(r,i_ucsb) else None),
                  asem=(str(g(r,i_asem)).strip() if g(r,i_asem) else None))
    role=str(g(r,i_pos)).strip() if g(r,i_pos) else "Member"
    appts.append(dict(pk=pk,bcu=bk,role=role,is_chair=("chair" in role.lower() and "vice" not in role.lower()),term="2025-26"))

# 4a. CoC budget
brows=list(load("2025-26 CoC Budget.xlsx").worksheets[0].iter_rows(values_only=True))
budget.append(dict(entity="CoC",fy="2025-26",cat="allocation",desc="Total CoC Budget",amt=g(brows[0],2),stage="actual",sort=0))
so=1
for r in brows[2:]:
    lab=g(r,0); val=g(r,2)
    if lab and val not in (None,""):
        labl=str(lab).strip(); low=labl.lower()
        cat="remaining" if "remaining" in low else ("honoraria" if "honoraria" in low else "expense")
        budget.append(dict(entity="CoC",fy="2025-26",cat=cat,desc=labl,amt=val,stage="actual",sort=so)); so+=1
    p5=g(r,5)
    if p5 and g(r,6) and str(p5).strip() not in ("Position","Total Honoraria:"):
        budget.append(dict(entity="CoC",fy="2025-26",cat="honoraria_cap",
                           desc=f"{str(p5).strip()} — {str(g(r,6)).strip()}",amt=g(r,7),stage="cap",sort=200+so)); so+=1

# 4b. AS FY27 per-BCU staff allocations
wb=load("26-27 AS UCSB Senate Budget Final.xlsx"); target=None
for ws in wb.worksheets:
    if "career staff" in ws.title.lower(): target=ws; break
if target:
    started=False; so=0
    for r in target.iter_rows(values_only=True):
        a=str(g(r,0)).strip() if g(r,0) else ""
        if a.lower()=="bcus": started=True; continue
        if not started: continue
        if a=="" : break
        val=g(r,1)
        if val in (None,""): continue
        budget.append(dict(entity=a,fy="2026-27",cat="staff_allocation",
                           desc="FY27 staff allocation (Senate)",amt=val,stage="senate",sort=so)); so+=1

# 4c. AS full Senate budget — spending by category ("where the money goes")
sen=None
for ws2 in wb.worksheets:
    if ws2.title.strip().lower()=="public - senate": sen=ws2; break
if sen:
    srows=[list(r) for r in sen.iter_rows(values_only=True)]
    hdr=srows[1]
    grp=[i for i,v in enumerate(hdr) if v and "RECOMMENDATION" not in str(v).upper() and i>=2]
    def snum(v):
        try: return float(v)
        except: return 0.0
    cat={}
    for r in srows:
        lab=str(g(r,0)).strip() if g(r,0) else ""
        if lab and lab[0].isdigit():
            cat[lab]=cat.get(lab,0.0)+sum(snum(g(r,ci+3)) for ci in grp)
    def pick(pred): return sum(v for k,v in cat.items() if pred(k))
    operating=pick(lambda k:(k.startswith("7000.") and not k.startswith("7000.00"))
                            or k.startswith("7001") or k.startswith("7800")
                            or k.startswith("7910") or k.startswith("8600"))
    buckets=[
        ("Career staff salaries", pick(lambda k:k.startswith("6200"))),
        ("Special & grant projects", pick(lambda k:k.startswith("7900") or k.startswith("7901"))),
        ("Operating expenses", operating),
        ("Contracts & contractors", pick(lambda k:k.startswith("7200") or k.startswith("6000"))),
        ("Student staff wages", pick(lambda k:k.startswith("6500") or k.startswith("6600"))),
        ("Grants", pick(lambda k:k.startswith("6700"))),
        ("Honoraria & stipends", pick(lambda k:k.startswith("7100") or k.startswith("6100"))),
    ]
    total=sum(a for _,a in buckets if a>0)
    budget.append(dict(entity="Associated Students",fy="2026-27",cat="as_total",
                       desc="Total AS spending (2026-27 Senate-recommended)",amt=round(total,2),stage="senate",sort=0))
    for i,(name,amt) in enumerate(sorted(buckets,key=lambda x:-x[1]),1):
        if amt<=0: continue
        budget.append(dict(entity="Associated Students",fy="2026-27",cat="as_category",
                           desc=name,amt=round(amt,2),stage="senate",sort=i))

# emit
L=["-- AUTO-GENERATED seed from /reference spreadsheets. Re-runnable on a fresh DB.","begin;",
   "truncate application_event, application, position, appointment, person, budget_line, bcu restart identity cascade;",""]
L.append("-- BCUs")
for b in bcus.values():
    L.append(f"insert into bcu(slug,name,short_name,type,website,contact_name,contact_email) values "
             f"({q(b['slug'])},{q(b['name'])},{q(b['short'])},'{b['type']}',{q(b['website'])},{q(b['cname'])},{q(b['cemail'])}) on conflict (slug) do nothing;")
L.append("\n-- People")
for p in persons.values():
    L.append(f"insert into person(full_name,ucsb_email,as_email) values ({q(p['full_name'])},{q(p['ucsb'])},{q(p['asem'])});")
def pexpr(pk):
    p=persons[pk]
    return (f"(select id from person where ucsb_email={q(p['ucsb'])} order by created_at limit 1)" if p['ucsb']
            else f"(select id from person where full_name={q(p['full_name'])} order by created_at limit 1)")
def bexpr(k): return f"(select id from bcu where slug={q(bcus[k]['slug'])} limit 1)"
L.append("\n-- Appointments")
for a in appts:
    L.append(f"insert into appointment(person_id,bcu_id,role_title,is_chair,term) values "
             f"({pexpr(a['pk'])},{bexpr(a['bcu'])},{q(a['role'])},{str(a['is_chair']).lower()},{q(a['term'])});")
L.append("\n-- Positions")
for p in positions:
    L.append(f"insert into position(bcu_id,title,status,routing,external_url,coc_advertises,notes) values "
             f"({bexpr(p['bcu'])},{q(p['title'])},'{p['status']}','{p['routing']}',{q(p['external_url'])},{str(p['coc_adv']).lower()},{q(p['notes'])});")
L.append("\n-- Budget")
for b in budget:
    L.append(f"insert into budget_line(entity,fiscal_year,category,description,amount,recommendation_stage,sort_order) values "
             f"({q(b['entity'])},{q(b['fy'])},{q(b['cat'])},{q(b['desc'])},{num(b['amt'])},{q(b['stage'])},{b['sort']});")
L.append("\ncommit;")
open(OUT,"w").write("\n".join(L)+"\n")
print(f"BCUs={len(bcus)} People={len(persons)} Appointments={len(appts)} Positions={len(positions)} BudgetLines={len(budget)}")
print("Wrote",OUT,os.path.getsize(OUT),"bytes")
